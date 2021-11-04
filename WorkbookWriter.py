from pathlib import Path
from openpyxl import load_workbook, Workbook 
from datetime import date
import calendar

class WorkbookWriter:
    def __init__(self,name):
        self.path = str(Path().resolve())+'/AttendanceSheets/'+name+'.xlsx'
        try:
            self.workbook = load_workbook(filename=self.path, read_only=False)
        except:
            self.create_new_sheet(name)
            self.workbook = load_workbook(filename=self.path, read_only=False)
        self.date = str(date.today())

    def fill_sheet(self, wb, name,fresh=True):
        month=calendar.month_name[int(str(date.today())[5:7])]
        max_days=calendar.monthrange(int(str(date.today())[:4]),int(str(date.today())[5:7]))[1]
        s=str(Path().resolve())+'/AttendanceSheets/'+name+'.xlsx'
        ws=None
        if fresh:
            ws=wb.active
            ws.title=month
        else:
            ws = wb.create_sheet(title=month, index=0)
        ws['A1']='Roll Number'
        ws['B1']='Name'
        hook=ws['B1']
        for i in range(0,max_days):
            hook.offset(row=0,column=i+1).value=i+1
        hook.offset(row=0,column=max_days+1).value='Total'
        for i in range(1,101):
            pass
        hook.offset(row=0,column=max_days+2).value='Percentage'
        wb.save(s)

    def create_new_sheet(self, name, wb=None):
        if wb==None:
            wb=Workbook()
            self.fill_sheet(wb,name)
        else:
            self.fill_sheet(wb,name, False)    

    def write(self,rollno,name,present):
        month=calendar.month_name[int(str(date.today())[5:7])]
        wb=self.workbook
        ws=wb[month]
        hook=ws['A1']
        row=int(rollno[-3:])
        col=int(self.date[-2:])+1
        hook.offset(row=row,column=0).value=rollno
        hook.offset(row=row,column=1).value=name
        if present:
            hook.offset(row=row,column=col).value='P'
        else:
            hook.offset(row=row,column=col).value='A'
        wb.save(self.path)